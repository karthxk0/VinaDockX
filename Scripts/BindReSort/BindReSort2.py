#!/usr/bin/env python3
import os
import sys
import glob
import re
import argparse
from datetime import datetime

# ==========================================
# Dependency Checking
# ==========================================
missing_deps = []
try:
    import pandas as pd
except ImportError:
    missing_deps.append("pandas")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    missing_deps.append("openpyxl")

try:
    from colorama import init, Fore, Style
except ImportError:
    missing_deps.append("colorama")

try:
    from tqdm import tqdm
except ImportError:
    missing_deps.append("tqdm")

if missing_deps:
    print("======================================================================")
    print(" ERROR: Missing required Python packages!")
    print(f" Please install them using: pip install {' '.join(missing_deps)}")
    print("======================================================================")
    sys.exit(1)

# Initialize colorama
init(autoreset=True)

VERSION = "2.2"

# ==========================================
# Banner
# ==========================================
def print_banner():
    banner = f"""
{Fore.LIGHTCYAN_EX}======================================================================                                                      
  ____  _           _ _____       _____            _   
 |  _ \\(_)         | |  __ \\     / ____|          | |  
 | |_) |_ _ __   __| | |__) |___| (___   ___  _ __| |_ 
 |  _ <| | '_ \\ / _` |  _  // _ \\\\___ \\ / _ \\| '__| __|
 | |_) | | | | | (_| | | \\ \\  __/____) | (_) | |  | |_ 
 |____/|_|_| |_|\\__,_|_|  \\_\\___|_____/ \\___/|_|   \\__|.py                                                     
                                                                        
 > Version {VERSION}  |  > Designed by karthxk (https://karthxk0.github.io/)         
  
======================================================================{Style.RESET_ALL}
"""
    print(banner)

# ==========================================
# Parsing Logic
# ==========================================
def extract_basename(path_str):
    if not path_str: return "Unknown"
    base = os.path.basename(path_str.strip())
    if base.lower().endswith('.pdbqt'):
        base = base[:-6]
    return base

def parse_vina_log(file_path):
    """Parses a single AutoDock Vina log file and yields dictionaries of ligand results."""
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read()

    blocks = re.split(r'={10,}|Reading input \.\.\. done\.', content)
    
    global_exhaust = re.search(r'Exhaustiveness:\s*(\d+)', content, re.IGNORECASE)
    global_exhaust = global_exhaust.group(1) if global_exhaust else "N/A"
    
    global_receptor = re.search(r'Rigid receptor:\s*([^\n\r]+)', content, re.IGNORECASE)
    global_receptor = global_receptor.group(1) if global_receptor else "Unknown"

    global_flex = re.search(r'Flex receptor:\s*([^\n\r]+)', content, re.IGNORECASE)
    global_flex = global_flex.group(1) if global_flex else None

    global_search_match = re.search(r'Search Space:\s*([^\n\r]+)', content, re.IGNORECASE)
    global_search = global_search_match.group(1).strip() if global_search_match else "N/A"

    results = []

    for block in blocks:
        if "mode |   affinity" not in block:
            continue
            
        receptor_match = re.search(r'Docking receptor\s+\'([^\']+)\'', block) or \
                         re.search(r'Rigid receptor:\s*([^\n\r]+)', block)
        receptor = extract_basename(receptor_match.group(1)) if receptor_match else extract_basename(global_receptor)

        ligand_match = re.search(r'with ligand\s+\'([^\']+)\'', block) or \
                       re.search(r'Ligand:\s*([^\n\r]+)', block)
        ligand = extract_basename(ligand_match.group(1)) if ligand_match else "Unknown_Ligand"

        flex_match = re.search(r'Flex receptor:\s*([^\n\r]+)', block)
        flex = extract_basename(flex_match.group(1)) if flex_match else (extract_basename(global_flex) if global_flex else None)

        exhaust_match = re.search(r'Exhaustiveness:\s*(\d+)', block)
        exhaust = exhaust_match.group(1) if exhaust_match else global_exhaust

        search_match = re.search(r'Search Space:\s*([^\n\r]+)', block, re.IGNORECASE)
        search_space = search_match.group(1).strip() if search_match else global_search

        seed_match = re.search(r'random seed:\s*([-\d]+)', block)
        seed = seed_match.group(1) if seed_match else "N/A"

        docking_mode = "Flexible Docking" if flex else "Standard Docking"

        table_lines = []
        capture = False
        for line in block.split('\n'):
            if "mode |   affinity" in line:
                capture = True
                table_lines.append(line)
                continue
            if capture:
                table_lines.append(line)
                if line.strip() == "" and len(table_lines) > 4:
                    break

        modes_data = []
        if len(table_lines) > 3:
            for line in table_lines[3:]:
                parts = line.split()
                if len(parts) >= 2 and parts[0].isdigit():
                    modes_data.append({
                        'mode': int(parts[0]),
                        'affinity': float(parts[1]),
                        'rmsd_lb': parts[2] if len(parts) > 2 else "0",
                        'rmsd_ub': parts[3] if len(parts) > 3 else "0"
                    })

        if modes_data:
            results.append({
                'receptor': receptor,
                'flex_receptor': flex,
                'ligand': ligand,
                'search_space': search_space,
                'docking_mode': docking_mode,
                'exhaustiveness': exhaust,
                'seed': seed,
                'table_raw': '\n'.join(table_lines).strip(),
                'modes': modes_data,
                'source': os.path.abspath(file_path)
            })

    return results

# ==========================================
# Output Generators
# ==========================================
def write_txt_output(output_path, ranked_data):
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"# Generated by BindReSort.py v{VERSION}\n")
        f.write("# Designed by karthxk (https://karthxk0.github.io/)\n\n")
        
        total = len(ranked_data)
        for i, data in enumerate(ranked_data):
            f.write("=================================================================\n\n")
            f.write(f"Rank: {i+1}/{total}\n\n")
            f.write(f"Ligand: {data['ligand']}\n")
            f.write(f"Receptor: {data['receptor']}\n")
            if data['flex_receptor']:
                f.write(f"Flex Receptor :  {data['flex_receptor']}\n")
            if data['search_space'] and data['search_space'] != "N/A":
                f.write(f"Search Space: {data['search_space']}\n")
            f.write(f"Docking Mode: {data['docking_mode']}\n")
            f.write(f"Exhaustiveness: {data['exhaustiveness']}\n")
            f.write(f"Random Seed: {data['seed']}\n\n")
            f.write("Results:\n\n")
            f.write(f"{data['table_raw']}\n\n")
            f.write(f"Source Log: {data['source']}\n\n")
        f.write("=================================================================\n")

def get_max_modes(data_list):
    max_mode = 0
    for d in data_list:
        for m in d['modes']:
            if m['mode'] > max_mode:
                max_mode = m['mode']
    return max_mode

def build_dataframe_for_mode(data_list, target_mode):
    rows = []
    
    # Check if any entry has a valid search space to decide if we need the column
    has_search_space = any(d['search_space'] and d['search_space'] != "N/A" for d in data_list)
    
    for d in data_list:
        for m in d['modes']:
            if m['mode'] == target_mode:
                row_dict = {
                    'Rank': 0, 
                    'Receptor': d['receptor'],
                    'Ligand': d['ligand'],
                    'Affinity (kcal/mol)': m['affinity']
                }
                
                if has_search_space:
                    if d['search_space'] != "N/A":
                        # Strip out anything inside parentheses
                        clean_space = re.sub(r'\s*\(.*?\)', '', d['search_space']).strip()
                        row_dict['Search Space'] = clean_space
                    else:
                        row_dict['Search Space'] = "N/A"
                        
                row_dict['Binding Mode'] = d['docking_mode'].split()[0]
                row_dict['Exhaustiveness'] = d['exhaustiveness']
                row_dict['Source Log'] = d['source']
                
                rows.append(row_dict)
    
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(by='Affinity (kcal/mol)', ascending=True).reset_index(drop=True)
        df['Rank'] = df.index + 1
    return df

def write_excel_output(output_path, data_list):
    max_mode = get_max_modes(data_list)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    for mode in range(1, max_mode + 1):
        df = build_dataframe_for_mode(data_list, mode)
        if df.empty: continue
        
        ws = wb.create_sheet(title=f"Mode {mode}")
        
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Generated by BindReSort.py v{VERSION}"
        ws['A1'].alignment = center_align
        
        ws.merge_cells('A2:G2')
        ws['A2'] = "Designed by karthxk (https://karthxk0.github.io/)"
        ws['A2'].alignment = center_align
        
        ws.merge_cells('A4:G4')
        ws['A4'] = f"Binding Mode: {mode}"
        ws['A4'].alignment = center_align
        
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = thin_border
            
        for row_idx, row_data in enumerate(df.values, 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align  
                cell.border = thin_border
        
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)
            
    wb.save(output_path)

def write_csv_outputs(output_dir, base_name, data_list):
    max_mode = get_max_modes(data_list)
    generated_files = []
    
    for mode in range(1, max_mode + 1):
        df = build_dataframe_for_mode(data_list, mode)
        if df.empty: continue
        
        csv_filename = f"{base_name}_mode{mode}.csv"
        csv_path = os.path.join(output_dir, csv_filename)
        
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write(f"Generated by BindReSort.py v{VERSION},,,,,,\n")
            f.write("Designed by karthxk (https://karthxk0.github.io/),,,,,,\n")
            f.write(",,,,,,\n")
            f.write(f"Binding Mode: {mode},,,,,,\n")
            
        df.to_csv(csv_path, mode='a', index=False)
        generated_files.append(csv_filename)
        
    return generated_files

# ==========================================
# File Gathering
# ==========================================
def get_files_from_input(input_str):
    paths = [p.strip(" \"'") for p in input_str.split(',')]
    txt_files = []
    
    for path in paths:
        if os.path.isfile(path) and path.endswith('.txt'):
            txt_files.append(path)
        elif os.path.isdir(path):
            search_pattern = os.path.join(path, '**', '*.txt')
            txt_files.extend(glob.glob(search_pattern, recursive=True))
        else:
            print(f"{Fore.LIGHTYELLOW_EX}[WARNING] Path not found or invalid: {path}")
            
    return list(set(txt_files))

# ==========================================
# Main Execution
# ==========================================
def main():
    parser = argparse.ArgumentParser(description="AutoDock Vina Log Sorter & Extractor")
    parser.add_argument("-i", "--input", help="Input files or directories (comma separated)", type=str)
    parser.add_argument("-o", "--output", help="Output directory", type=str)
    parser.add_argument("-f", "--format", help="Output format: 1 (CSV), 2 (Excel), 3 (Both)", type=int, choices=[1, 2, 3])
    
    args = parser.parse_args()
    print_banner()

    if not args.input:
        print(f"{Fore.LIGHTYELLOW_EX}Enter VinaDock Log (.txt) files or directories (separated by commas):")
        inputs = input(f"{Fore.LIGHTCYAN_EX}> ").strip()
    else:
        inputs = args.input

    if not args.output:
        print(f"\n{Fore.LIGHTYELLOW_EX}Enter the directory where you wish to save the output:")
        out_dir = input(f"{Fore.LIGHTCYAN_EX}> ").strip(" \"'")
    else:
        out_dir = args.output.strip(" \"'")

    if not args.format:
        print(f"\n{Fore.LIGHTYELLOW_EX}In which format do you wish the sorted output to be?")
        print(f"{Fore.LIGHTYELLOW_EX}1. CSV File")
        print(f"{Fore.LIGHTYELLOW_EX}2. Excel (XLSX) File")
        print(f"{Fore.LIGHTYELLOW_EX}3. Both")
        while True:
            try:
                fmt = int(input(f"{Fore.LIGHTCYAN_EX}> ").strip())
                if fmt in [1, 2, 3]: break
                print(f"{Fore.LIGHTRED_EX}Please enter 1, 2, or 3.")
            except ValueError:
                print(f"{Fore.LIGHTRED_EX}Invalid input. Please enter a number.")
    else:
        fmt = args.format

    files_to_process = get_files_from_input(inputs)
    if not files_to_process:
        print(f"{Fore.LIGHTRED_EX}\n[ERROR] No valid .txt log files found. Exiting.")
        sys.exit(1)

    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
        print(f"{Fore.LIGHTBLUE_EX}[INFO] Created output directory: {out_dir}")

    print(f"\n{Fore.LIGHTMAGENTA_EX}================== PROCESSING LOGS ==================")
    all_data = []
    failed_files = []

    for file in tqdm(files_to_process, desc=f"{Fore.LIGHTCYAN_EX}Parsing files{Style.RESET_ALL}", bar_format='{l_bar}{bar:30}{r_bar}'):
        try:
            parsed = parse_vina_log(file)
            if parsed:
                all_data.extend(parsed)
            else:
                failed_files.append((file, "No valid table/data found"))
        except Exception as e:
            failed_files.append((file, str(e)))

    if not all_data:
        print(f"{Fore.LIGHTRED_EX}\n[ERROR] Could not extract any valid docking data from the provided files.")
        sys.exit(1)

    all_data.sort(key=lambda x: x['modes'][0]['affinity'] if x['modes'] else 999)

    unique_receptors = list(set([d['receptor'] for d in all_data]))
    receptors_str = "-".join(unique_receptors[:3]) 
    if len(unique_receptors) > 3:
        receptors_str += "-etc"
    
    timestamp = datetime.now().strftime("%d%m%Y-%H%M")
    base_filename = f"{receptors_str}_BindReSort_{timestamp}"
    
    print(f"\n{Fore.LIGHTMAGENTA_EX}================== GENERATING OUTPUT ==================")
    
    txt_path = os.path.join(out_dir, f"{base_filename}.txt")
    write_txt_output(txt_path, all_data)
    print(f"{Fore.LIGHTGREEN_EX}[+] Created Text Log: {Fore.LIGHTWHITE_EX}{os.path.basename(txt_path)}")
    
    generated_files = 1

    if fmt in [2, 3]:
        xl_path = os.path.join(out_dir, f"{base_filename}.xlsx")
        write_excel_output(xl_path, all_data)
        print(f"{Fore.LIGHTGREEN_EX}[+] Created Excel File: {Fore.LIGHTWHITE_EX}{os.path.basename(xl_path)}")
        generated_files += 1

    if fmt in [1, 3]:
        csv_files = write_csv_outputs(out_dir, base_filename, all_data)
        for cf in csv_files:
            print(f"{Fore.LIGHTGREEN_EX}[+] Created CSV File: {Fore.LIGHTWHITE_EX}{cf}")
            generated_files += 1

    print(f"\n{Fore.LIGHTCYAN_EX}======================================================================")
    print(f"{Fore.LIGHTYELLOW_EX}                        SORTING SUMMARY")
    print(f"{Fore.LIGHTCYAN_EX}======================================================================")
    print(f"{Fore.LIGHTWHITE_EX}Total Input Files Scanned   : {Fore.LIGHTCYAN_EX}{len(files_to_process)}")
    print(f"{Fore.LIGHTWHITE_EX}Total Valid Ligands Found   : {Fore.LIGHTGREEN_EX}{len(all_data)}")
    print(f"{Fore.LIGHTWHITE_EX}Unique Receptors Processed  : {Fore.LIGHTCYAN_EX}{len(unique_receptors)} {unique_receptors}")
    print(f"{Fore.LIGHTWHITE_EX}Total Output Files Generated: {Fore.LIGHTGREEN_EX}{generated_files}")
    
    if failed_files:
        print(f"\n{Fore.LIGHTRED_EX}Failures / Skipped Files: {len(failed_files)}")
        for f, err in failed_files:
            print(f"  - {Fore.LIGHTRED_EX}{os.path.basename(f)}{Style.RESET_ALL} (Reason: {err})")
    else:
        print(f"\n{Fore.LIGHTGREEN_EX}All files processed successfully! 0 Failures.")

    print(f"\n{Fore.LIGHTGREEN_EX}✔ Task Completed! Your sorted results are saved in:{Style.RESET_ALL}")
    print(f"{Fore.LIGHTWHITE_EX}{os.path.abspath(out_dir)}")
    print(f"{Fore.LIGHTCYAN_EX}======================================================================\n")

if __name__ == "__main__":
    main()